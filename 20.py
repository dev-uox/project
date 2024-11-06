import pandas as pd
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import threading
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

class ExcelEditor:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Search Tool for Everyone")

        # Set the window size to be responsive
        self.root.geometry("900x600")

        self.selected_row_index = None  # Track selected row index

        # Load the Excel file in a separate thread
        threading.Thread(target=self.load_excel_in_thread).start()

    def load_excel_in_thread(self):
        # File path provided here
        self.file_path = "C:/Users/Admin/Downloads/Spectrum Confirmations 2024.xlsx"
        if self.file_path:
            self.load_excel()
            self.create_widgets()
        else:
            messagebox.showerror("Error", "No file selected!")
            self.root.quit()

    def load_excel(self):
        try:
            self.df = pd.read_excel(self.file_path)
            self.df.reset_index(inplace=True)  # Reset index to use row numbers
            self.df.rename(columns={'index': 'Row Number'}, inplace=True)  # Rename the index column to "Row Number"
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load Excel file: {e}")
            self.root.quit()

    def create_widgets(self):
        self.root.after(0, self._create_widgets)

    def _create_widgets(self):
        tk.Label(self.root, text="Full Excel Data:").pack(padx=10, pady=10)

        # Create a frame to hold the Treeview and scrollbars
        frame = tk.Frame(self.root)
        frame.pack(fill=tk.BOTH, expand=True)

        # Create a Treeview to display the DataFrame with horizontal and vertical scrollbars
        self.columns = ['Row Number'] + self.df.columns.tolist()[1:]  # Include Row Number
        self.tree = ttk.Treeview(frame, columns=self.columns, show="headings")

        # Add scrollbars
        vsb = ttk.Scrollbar(frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        hsb.pack(side=tk.BOTTOM, fill=tk.X)

        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Set up columns and headings
        for col in self.columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=100, anchor='center')

        # Insert data into the Treeview with alternating row colors for visual separation
        for idx, row in self.df.iterrows():
            if idx % 2 == 0:
                self.tree.insert("", "end", values=row.tolist(), tags=('evenrow',))
            else:
                self.tree.insert("", "end", values=row.tolist(), tags=('oddrow',))

        # Add row coloring
        self.tree.tag_configure('evenrow', background='lightblue')  # Even rows
        self.tree.tag_configure('oddrow', background='lightgray')  # Odd rows

        # Filter widgets
        filter_frame = tk.Frame(self.root)
        filter_frame.pack(pady=10)

        tk.Label(filter_frame, text="Filter by Column:").grid(row=0, column=0, padx=5)

        # Dropdown to select the column for filtering
        self.selected_column = tk.StringVar()
        self.selected_column.set(self.columns[1])  # Default to first data column
        column_dropdown = ttk.OptionMenu(filter_frame, self.selected_column, *self.columns[1:])
        column_dropdown.grid(row=0, column=1, padx=5)

        # Entry box to input filter value
        tk.Label(filter_frame, text="Filter Value:").grid(row=0, column=2, padx=5)
        self.filter_value_entry = tk.Entry(filter_frame)
        self.filter_value_entry.grid(row=0, column=3, padx=5)

        # Button to filter the data
        tk.Button(filter_frame, text="Filter", command=self.filter_data).grid(row=0, column=4, padx=5)

        # Button to clear filter and show all data
        tk.Button(filter_frame, text="Clear Filter", command=self.clear_filter).grid(row=0, column=5, padx=5)

        # Enable cell editing when a user double-clicks a cell
        self.tree.bind("<Double-1>", self.on_double_click)

        tk.Button(self.root, text="Save Changes", command=self.save_changes).pack(padx=10, pady=10)

    def filter_data(self):
        filter_value = self.filter_value_entry.get()
        selected_column = self.selected_column.get()

        if not filter_value:
            messagebox.showerror("Error", "Please enter a filter value.")
            return

        # Filter the DataFrame based on user input
        filtered_df = self.df[self.df[selected_column].astype(str).str.contains(filter_value, na=False, case=False)]

        # Clear the Treeview
        self.tree.delete(*self.tree.get_children())

        # Insert filtered data into the Treeview with alternating colors
        for idx, row in filtered_df.iterrows():
            if idx % 2 == 0:
                self.tree.insert("", "end", values=row.tolist(), tags=('evenrow',))
            else:
                self.tree.insert("", "end", values=row.tolist(), tags=('oddrow',))

    def clear_filter(self):
        self.filter_value_entry.delete(0, tk.END)  # Clear the entry box

        # Clear the Treeview
        self.tree.delete(*self.tree.get_children())

        # Insert all data into the Treeview with alternating colors
        for idx, row in self.df.iterrows():
            if idx % 2 == 0:
                self.tree.insert("", "end", values=row.tolist(), tags=('evenrow',))
            else:
                self.tree.insert("", "end", values=row.tolist(), tags=('oddrow',))

    def on_double_click(self, event):
        # Get selected cell
        item = self.tree.selection()[0]
        column = self.tree.identify_column(event.x)  # Get column number
        row = self.tree.identify_row(event.y)  # Get row number

        col_idx = int(column.replace('#', '')) - 1  # Convert from string to int (1-indexed)
        row_idx = self.tree.index(item)  # Get the index of the row

        # Get current value in the selected cell
        current_value = self.tree.item(item, 'values')[col_idx]

        # Create an enlarged editing window at the cell location
        self.edit_popup = tk.Toplevel(self.root)
        self.edit_popup.title("Edit Cell Value")
        self.edit_popup.geometry(f"400x200+{event.x_root}+{event.y_root}")

        # Label for current value
        tk.Label(self.edit_popup, text="Current Value:").pack(pady=5)
        tk.Label(self.edit_popup, text=current_value, wraplength=350).pack(pady=5)

        # Entry widget to allow the user to edit the value
        tk.Label(self.edit_popup, text="New Value:").pack(pady=5)
        entry = tk.Entry(self.edit_popup, width=50)  # Make the Entry widget wider for easier editing
        entry.insert(0, current_value)
        entry.pack(pady=10)

        # Button to save the updated value
        save_button = tk.Button(self.edit_popup, text="Save", command=lambda: self.save_edit(row_idx, col_idx, entry.get()))
        save_button.pack(pady=10)

    def save_edit(self, row_idx, col_idx, value):
        # Update the value in the DataFrame
        col_name = self.df.columns[col_idx]
        self.df.at[row_idx, col_name] = value

        # Update the Treeview with the new value
        self.tree.set(self.tree.get_children()[row_idx], column=col_idx, value=value)

        # Destroy the edit popup window after saving the changes
        self.edit_popup.destroy()


    def save_edit(self, row_idx, col_idx, value):
        # Update the value in the DataFrame
        col_name = self.df.columns[col_idx]
        self.df.at[row_idx, col_name] = value

        # Update the Treeview with the new value
        self.tree.set(self.tree.get_children()[row_idx], column=col_idx, value=value)

        self.edit_popup.destroy()

    def save_changes(self):
        try:
            # Load the existing Excel file with openpyxl
            workbook = load_workbook(self.file_path)
            sheet = workbook.active

            # Clear the existing data in the sheet
            sheet.delete_rows(2, sheet.max_row)

            # Write the updated DataFrame back to the Excel file, starting from the second row
            for r_idx, row in enumerate(dataframe_to_rows(self.df, index=False, header=False), 2):
                for c_idx, value in enumerate(row, 1):
                    sheet.cell(row=r_idx, column=c_idx, value=value)

            # Save the workbook after updating it
            workbook.save(self.file_path)

            messagebox.showinfo("Success", "Changes saved to Excel file, and formatting is preserved.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save changes: {e}")

# Create the Tkinter app
root = tk.Tk()

app = ExcelEditor(root)
root.mainloop()
