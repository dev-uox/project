import pandas as pd
import tkinter as tk
from tkinter import messagebox, filedialog, Toplevel
import threading
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

class ExcelEditor:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Search Tool for Everyone")
        
        self.is_case_sensitive = tk.BooleanVar(value=False)
        self.selected_row_index = None  # Track selected row index

        # Load the Excel file in a separate thread
        threading.Thread(target=self.load_excel_in_thread).start()

    def load_excel_in_thread(self):
        self.file_path = filedialog.askopenfilename(title="Select Excel File", filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")))
        if self.file_path:
            self.load_excel()
            self.create_widgets()
        else:
            messagebox.showerror("Error", "No file selected!")
            self.root.quit()

    def load_excel(self):
        try:
            self.df = pd.read_excel(self.file_path)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load Excel file: {e}")
            self.root.quit()

    def create_widgets(self):
        self.root.after(0, self._create_widgets)

    def _create_widgets(self):
        tk.Label(self.root, text="Search by:").grid(row=0, column=0, padx=10, pady=10)

        self.search_column = tk.StringVar(value="Business Name")
        tk.OptionMenu(self.root, self.search_column, "Business Name (What is the company called?)", 
                      "Address (Where is the company located?)", 
                      "Cell No (What's the phone number?)",
                      "Email").grid(row=0, column=1)

        self.search_entry = tk.Entry(self.root)
        self.search_entry.grid(row=0, column=2, padx=10, pady=10)
        self.search_entry.insert(0, "Type here...")

        tk.Button(self.root, text="Search", command=self.search_data).grid(row=0, column=3, padx=10, pady=10)

        tk.Checkbutton(self.root, text="Case Sensitive", variable=self.is_case_sensitive).grid(row=0, column=4, padx=10, pady=10)

        self.text_box = tk.Text(self.root, width=100, height=20, wrap=tk.WORD, bg="lightyellow")
        self.text_box.grid(row=1, column=0, columnspan=5, padx=10, pady=10)

        scrollbar = tk.Scrollbar(self.root, command=self.text_box.yview)
        scrollbar.grid(row=1, column=5, sticky='ns')
        self.text_box['yscrollcommand'] = scrollbar.set

        tk.Button(self.root, text="Clear Results", command=self.clear_results).grid(row=2, column=1, padx=10, pady=10)
        tk.Button(self.root, text="Help", command=self.show_help).grid(row=2, column=2, padx=10, pady=10)
        tk.Button(self.root, text="Edit Selected", command=self.edit_selected).grid(row=2, column=3, padx=10, pady=10)
        tk.Button(self.root, text="Save Changes", command=self.save_changes).grid(row=2, column=4, padx=10, pady=10)
        tk.Button(self.root, text="Send Emails", command=self.send_emails).grid(row=2, column=0, padx=10, pady=10)

    def clear_results(self):
        self.text_box.delete(1.0, tk.END)

    def show_help(self):
        help_message = (
            "How to Use the Search Tool:\n"
            "1. Choose what you want to search for (Business Name, Address, Cell No, or Email).\n"
            "2. Type in what you're looking for.\n"
            "3. Click the 'Search' button to see the results.\n"
            "4. You can clear the results anytime by clicking 'Clear Results'.\n"
            "5. Select a result and click 'Edit Selected' to modify it.\n"
            "6. Click 'Save Changes' to update the Excel file."
        )
        messagebox.showinfo("Help", help_message)

    def format_results(self, results):
        formatted_text = "Here are the details for your search:\n\n"
        for idx, row in results.iterrows():
            formatted_text += "-------------------------------------------------------\n"
            formatted_text += f"Index: {idx}\n"
            formatted_text += f"Business Name: {row['Business Name']}\n"
            formatted_text += f"Address: {row.get('Address', 'N/A')}\n"
            formatted_text += f"Cell No: {row.get('Cell No', 'N/A')}\n"
            formatted_text += f"Email: {row.get('Email', 'N/A')}\n"
            formatted_text += "-------------------------------------------------------\n\n"
        return formatted_text
    
    def search_data(self):
        search_value = self.search_entry.get()
        search_column = self.search_column.get()

        if not search_value:
            messagebox.showerror("Error", "Please enter a search value.")
            return

        column_map = {
            "Business Name (What is the company called?)": "Business Name",
            "Address (Where is the company located?)": "Address",
            "Cell No (What's the phone number?)": "Cell No",
            "Email": "Email"
        }
        search_column = column_map[search_column]

        if self.is_case_sensitive.get():
            results = self.df[self.df[search_column].astype(str).str.contains(search_value, na=False)]
        else:
            results = self.df[self.df[search_column].astype(str).str.contains(search_value, case=False, na=False)]

        if results.empty:
            self.text_box.delete(1.0, tk.END)
            self.text_box.insert(tk.END, "Oops! No results found.")
        else:
            self.text_box.delete(1.0, tk.END)
            formatted_results = self.format_results(results)
            self.text_box.insert(tk.END, formatted_results)

    def edit_selected(self):
        selected_text = self.text_box.get("1.0", tk.END).strip().split("\n")
        if not selected_text or "Oops! No results found." in selected_text[0]:
            messagebox.showerror("Error", "No results to edit.")
            return

        # Assume the first result is being edited. This can be enhanced for user selection.
        index_line = selected_text[1]  # The line with the index
        self.selected_row_index = int(index_line.split(":")[1].strip())

        # Open edit window
        edit_window = tk.Toplevel(self.root)
        edit_window.title("Edit Entry")

        self.edit_vars = {}
        for key in self.df.columns:
            value = self.df.at[self.selected_row_index, key]
            tk.Label(edit_window, text=key).grid(row=len(self.edit_vars), column=0)
            entry = tk.Entry(edit_window)
            entry.grid(row=len(self.edit_vars), column=1)
            entry.insert(0, value)
            self.edit_vars[key] = entry

        tk.Button(edit_window, text="Save", command=lambda: self.save_edit(edit_window)).grid(row=len(self.edit_vars), column=0, columnspan=2)

    def save_edit(self, edit_window):
        for key, entry in self.edit_vars.items():
            value = entry.get()
            self.df.at[self.selected_row_index, key] = value  # Update the DataFrame

        edit_window.destroy()
        messagebox.showinfo("Success", "Entry updated. Remember to save changes.")

    def save_changes(self):
        try:
            # Load the existing Excel file with openpyxl
            workbook = load_workbook(self.file_path)
            sheet = workbook.active

            # Clear the existing data in the sheet
            sheet.delete_rows(2, sheet.max_row)

            # Write the updated DataFrame back to the Excel file, starting from the second row
            for r_idx, row in enumerate(dataframe_to_rows(self.df, index=False, header=False), 2):
                for c_idx, value in enumerate(row, 1):
                    sheet.cell(row=r_idx, column=c_idx, value=value)

            # Save the workbook after updating it
            workbook.save(self.file_path)

            messagebox.showinfo("Success", "Changes saved to Excel file, and formatting is preserved.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save changes: {e}")

    def send_emails(self):
        email_column = "Email"
        if email_column not in self.df.columns:
            messagebox.showerror("Error", "No email column found in the Excel file.")
            return

        self.recipients_df = self.df.dropna(subset=[email_column])  # Drop rows where email is missing

        if self.recipients_df.empty:
            messagebox.showerror("Error", "No email addresses found.")
            return

        self.current_recipient_index = 0  # Keep track of the current recipient
        self.show_confirmation_window()  # Open the confirmation window

    def show_confirmation_window(self):
        if self.current_recipient_index >= len(self.recipients_df):
            messagebox.showinfo("Success", "Emails sent successfully!")
            return

        recipient_row = self.recipients_df.iloc[self.current_recipient_index]  # Get the current recipient's row

        recipient_email = recipient_row['Email']
        recipient_name = recipient_row['Business Name']
        address = recipient_row['Address']
        plan_details = recipient_row['Plan Details']

        # Create a new window for confirmation
        self.confirm_window = Toplevel(self.root)
        self.confirm_window.title(f"Confirm Email for {recipient_name}")

        tk.Label(self.confirm_window, text=f"Recipient Email:").grid(row=0, column=0)
        self.email_entry = tk.Entry(self.confirm_window, width=50)
        self.email_entry.grid(row=0, column=1)
        self.email_entry.insert(0, recipient_email)

        tk.Label(self.confirm_window, text=f"Subject:").grid(row=1, column=0)
        self.subject_entry = tk.Entry(self.confirm_window, width=50)
        self.subject_entry.grid(row=1, column=1)
        self.subject_entry.insert(0, "Your Business Proposal")

        tk.Label(self.confirm_window, text=f"Body:").grid(row=2, column=0)
        self.body_text = tk.Text(self.confirm_window, width=50, height=10)
        self.body_text.grid(row=2, column=1)

        # Define the template
        email_template = f"""
        Dear {recipient_name},

        This is Harvey from Spectrum Business. {address}. This address is Serviceable; Proceed with an Installation Truck Roll.

        YOUR SUGGESTED SPECTRUM BUSINESS PLAN:
        {plan_details}

        Note: To get a $5 discount on your total bill, please set up the autopay after your first bill.

        -Includes one FREE line from Spectrum Mobile™ for 12 months
        -No Contracts, Hidden Fees, or Added Phone Taxes
        -Lock Your Price for up to 3 years
        -Over 99.9% Network Reliability
        -Includes FREE Internet features (over $50/mo value)

        Business Voice Features: 35+ business-centric phone features like Voicemail, Custom Caller ID, Hunt Groups, Account Codes, and Three-Way Calling are included at no additional charge. Nationwide Unlimited Calling is included for FREE.

        Thank you in advance for your time. Have a great day ahead!

        Best regards,
        Harvey Ham
        Sales & Marketing
        """

        self.body_text.insert(tk.END, email_template)

        tk.Button(self.confirm_window, text="Send", command=self.send_current_email).grid(row=3, column=1)
        tk.Button(self.confirm_window, text="Skip", command=self.skip_current_email).grid(row=3, column=0)

    def send_current_email(self):
        recipient_email = self.email_entry.get()
        subject = self.subject_entry.get()
        body = self.body_text.get("1.0", tk.END)

        # Email credentials
        sender_email = "developervahlayconsulting@gmail.com"
        sender_password = "vzlf yeqs tawu oynn"  # App-specific password

        try:
            msg = MIMEMultipart()
            msg['From'] = sender_email
            msg['To'] = recipient_email
            msg['Subject'] = subject
            msg.attach(MIMEText(body, 'plain'))

            # Connect to the SMTP server
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login(sender_email, sender_password)

            # Send the email
            server.sendmail(sender_email, recipient_email, msg.as_string())
            server.quit()
            print(f"Email sent to {recipient_email}")
        except Exception as e:
            print(f"Failed to send email to {recipient_email}: {e}")

        # Close the confirmation window and move to the next recipient
        self.confirm_window.destroy()
        self.current_recipient_index += 1
        self.show_confirmation_window()

    def skip_current_email(self):
        # Just close the confirmation window and move to the next recipient
        self.confirm_window.destroy()
        self.current_recipient_index += 1
        self.show_confirmation_window()

# Create the Tkinter app
root = tk.Tk()
app = ExcelEditor(root)
root.mainloop()
